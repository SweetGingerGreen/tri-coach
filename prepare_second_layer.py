#!/usr/bin/env python3
"""
Prepare second-layer reference docs and bike approved tool cards.

This is a one-time, idempotent ingest helper. It keeps original inbox files in
place, writes curated Markdown copies into the RAG folders, and adds consistent
frontmatter for chunk_approved.py.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).parent.resolve()
KNOWLEDGE = ROOT / "triathlon-knowledge"
INBOX = KNOWLEDGE / "00_inbox"
APPROVED = KNOWLEDGE / "01_approved"
REFERENCE = KNOWLEDGE / "02_reference"
TODAY = "2026-05-03"
COLLECTION = "triathlon_core_v2"


@dataclass(frozen=True)
class ReferenceItem:
    domain: str
    trust_level: str
    source_path: str
    summary: str
    usage_rule: str
    tags: list[str]


REFERENCE_ITEMS = [
    ReferenceItem(
        domain="nutrition",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/运动营养学/高级运动营养学（第2版） (丹·贝纳多特).md",
        summary="第二层运动营养参考，补充能量、碳水、蛋白质、液体和补剂等主题。",
        usage_rule="作为营养二线参考使用；具体克数建议必须结合体重、训练时长、强度、胃肠耐受和既往实践。",
        tags=["nutrition", "energy", "carbohydrate", "hydration", "sports_nutrition"],
    ),
    ReferenceItem(
        domain="nutrition",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/运动营养学/中国居民膳食指南（2022） (中国营养学会)/auto/中国居民膳食指南（2022） (中国营养学会).md",
        summary="公共膳食指南参考，用于日常饮食结构和健康饮食底线。",
        usage_rule="用于基础饮食结构参考；不能替代运动专项补给资料，也不能生成缺少个人数据的训练补给处方。",
        tags=["nutrition", "dietary_guidelines", "daily_diet", "health"],
    ),
    ReferenceItem(
        domain="run",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/跑步/悦动空间·跑步训练 汉森马拉松训练法 (（美）卢克·汉弗莱，凯斯·汉森，凯文·汉斯著；王晓刚译)/auto/悦动空间·跑步训练 汉森马拉松训练法 (（美）卢克·汉弗莱，凯斯·汉森，凯文·汉斯著；王晓刚译).md",
        summary="马拉松训练方法参考，可补充跑步有氧、节奏跑和长期计划思路。",
        usage_rule="只能作为跑步专项二线参考；铁三计划必须考虑骑跑疲劳叠加，不能直接套用马拉松课表。",
        tags=["run", "marathon", "hanson_method", "tempo", "aerobic"],
    ),
    ReferenceItem(
        domain="run",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/跑步/刷新PB：跑步提速指南 (【美】霍尔·希格登  谢维译).md",
        summary="大众跑步提速参考，适合补充跑步训练组织和比赛准备。",
        usage_rule="作为跑步补充资料使用；强度安排仍需优先服从铁三总负荷和恢复状态。",
        tags=["run", "speed", "race_preparation", "training_plan"],
    ),
    ReferenceItem(
        domain="run",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/跑步/无伤跑法 (戴剑松, 郑家轩).md",
        summary="本土跑步技术和伤病预防参考，可辅助识别常见跑步风险。",
        usage_rule="作为跑步技术与风险识别参考；不得诊断伤病，不得替代医生或物理治疗师。",
        tags=["run", "injury_prevention", "running_form", "risk_control"],
    ),
    ReferenceItem(
        domain="run",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/跑步/无伤跑法2：跑步技术优化与训练提升 (戴剑松).md",
        summary="跑步技术优化与训练提升参考，可与核心跑步资料交叉使用。",
        usage_rule="作为跑步技术与训练提升参考；遇到疼痛、麻木、无力或进行性加重时必须建议就医评估。",
        tags=["run", "running_form", "injury_prevention", "training_improvement"],
    ),
    ReferenceItem(
        domain="swim",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/游泳/The Swimming Drill Book, 2E (Guzman, Ruben).md",
        summary="游泳 drill 参考，适合后续拆成动作卡片库。",
        usage_rule="作为游泳动作练习补充；具体动作选择要结合当前技术短板，不要一次堆叠过多 drill。",
        tags=["swim", "drills", "technique", "skill_practice"],
    ),
    ReferenceItem(
        domain="swim",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/游泳/力争上游 100个游泳技巧完全图解=THE 100 BEST SWIMMING DRILLS (（美）布莱斯卢塞罗（BLYTHE LUCERO）著)/auto/力争上游 100个游泳技巧完全图解=THE 100 BEST SWIMMING DRILLS (（美）布莱斯卢塞罗（BLYTHE LUCERO）著).md",
        summary="游泳技巧图解和 drill 参考，适合做技术动作库来源。",
        usage_rule="作为游泳技术练习参考；需要和核心游泳训练资料交叉验证训练量和强度。",
        tags=["swim", "drills", "technique", "visual_drills"],
    ),
    ReferenceItem(
        domain="swim",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/游泳/游泳运动系统训练：运动原理、肌肉训练、运动损伤的预防 (伊恩·麦克劳德 (Ian Mcleod))/auto/游泳运动系统训练：运动原理、肌肉训练、运动损伤的预防 (伊恩·麦克劳德 (Ian Mcleod)).md",
        summary="游泳专项体能、运动原理和伤病预防参考。",
        usage_rule="作为游泳专项体能二线参考；康复和损伤内容只做风险提示，不能输出治疗处方。",
        tags=["swim", "strength", "injury_prevention", "mobility"],
    ),
    ReferenceItem(
        domain="strength",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/训练学/Triphasic Training (Cal Dietz Ben Peterson).md",
        summary="三相训练高阶力量周期化参考，适合高级力量专题。",
        usage_rule="作为高阶力量训练参考；普通铁三课表不得直接套用高冲击、高负荷方案。",
        tags=["strength", "triphasic_training", "periodization", "power"],
    ),
    ReferenceItem(
        domain="strength",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/训练学/Triphasic Training II (Cal Dietz Mike T Nelson) 三相训练 2.md",
        summary="三相训练二册参考，补充力量周期化和高阶体能安排。",
        usage_rule="作为高阶力量训练参考；必须结合训练年龄、恢复能力和铁三总负荷谨慎使用。",
        tags=["strength", "triphasic_training", "periodization", "power"],
    ),
    ReferenceItem(
        domain="strength",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/解剖按摩损伤/核心评估与训练：核心能力的精准测试与针对性发展（修订版） (美国人体运动出版社).md",
        summary="核心能力评估和训练参考，适合做动作评估和辅助训练来源。",
        usage_rule="作为核心评估与训练参考；动作选择要服务于铁三训练，不替代康复处方。",
        tags=["strength", "core", "assessment", "movement"],
    ),
    ReferenceItem(
        domain="recovery",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/解剖按摩损伤/人体解剖学彩色图谱·运动系统 (沃纳·普拉策(Werner Platzer))/auto/人体解剖学彩色图谱·运动系统 (沃纳·普拉策(Werner Platzer)).md",
        summary="运动系统解剖图谱参考，用于术语、结构和位置解释。",
        usage_rule="仅用于解剖背景和术语解释；不能据此诊断或制定治疗方案。",
        tags=["recovery", "anatomy", "movement_system", "terminology"],
    ),
    ReferenceItem(
        domain="recovery",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/解剖按摩损伤/运动按摩(肌肉训练彩色解剖图谱) (阿比盖尔·埃尔斯沃思, 佩姬·奥尔特曼著)/auto/运动按摩(肌肉训练彩色解剖图谱) (阿比盖尔·埃尔斯沃思, 佩姬·奥尔特曼著).md",
        summary="运动按摩和自我护理参考，用于恢复和放松背景知识。",
        usage_rule="仅作为恢复和放松参考；疼痛、麻木、无力或症状加重时必须建议专业评估。",
        tags=["recovery", "massage", "self_care", "mobility"],
    ),
    ReferenceItem(
        domain="race",
        trust_level="B",
        source_path="triathlon-knowledge/00_inbox/processed/Triathlete Magazines Complete Triathlon Book The Training, Diet, Health, Equipment, and Safety Tips You Need to Do Your Best (Matt Fitzgerald)/ocr/Triathlete Magazines Complete Triathlon Book The Training, Diet, Health, Equipment, and Safety Tips You Need to Do Your Best (Matt Fitzgerald).md",
        summary="铁三综合参考书 OCR 可读版本，覆盖装备、骑行、跑步、训练组织、营养和比赛执行。",
        usage_rule="作为第二层综合参考使用；其中 Bike Training 专章已单独抽入 bike approved，问答时优先采用 approved bike 卡片。",
        tags=["triathlon", "race", "bike", "run", "swim", "equipment", "nutrition"],
    ),
]


def rel(path: Path) -> str:
    return str(path.resolve().relative_to(ROOT))


def strip_frontmatter(text: str) -> str:
    if not text.startswith("---"):
        return text.lstrip()
    match = re.match(r"^---\s*\n.*?\n---\s*\n?", text, flags=re.DOTALL)
    if not match:
        return text.lstrip()
    return text[match.end() :].lstrip()


def parse_source_frontmatter(text: str) -> dict[str, str]:
    if not text.startswith("---"):
        return {}
    match = re.match(r"^---\s*\n(.*?)\n---\s*\n?", text, flags=re.DOTALL)
    if not match:
        return {}
    metadata: dict[str, str] = {}
    for line in match.group(1).splitlines():
        if ":" not in line or line.startswith(" "):
            continue
        key, value = line.split(":", 1)
        metadata[key.strip()] = value.strip().strip("\"'")
    return metadata


def yaml_value(value: Any) -> str:
    if isinstance(value, list):
        return json.dumps(value, ensure_ascii=False)
    return json.dumps(str(value), ensure_ascii=False)


def frontmatter(metadata: dict[str, Any]) -> str:
    lines = ["---"]
    for key, value in metadata.items():
        lines.append(f"{key}: {yaml_value(value)}")
    lines.append("---")
    return "\n".join(lines) + "\n\n"


def write_markdown(path: Path, metadata: dict[str, Any], body: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(frontmatter(metadata) + body.rstrip() + "\n", encoding="utf-8")


def infer_title(path: Path, source_metadata: dict[str, str]) -> str:
    return source_metadata.get("title") or path.stem


def reference_destination(item: ReferenceItem, source: Path) -> Path:
    return REFERENCE / item.domain / source.name


def prepare_references() -> list[Path]:
    written: list[Path] = []
    for item in REFERENCE_ITEMS:
        source = ROOT / item.source_path
        if not source.exists():
            raise FileNotFoundError(source)
        raw = source.read_text(encoding="utf-8", errors="ignore")
        source_meta = parse_source_frontmatter(raw)
        metadata = {
            "title": infer_title(source, source_meta),
            "domain": item.domain,
            "source": source_meta.get("source", item.source_path),
            "trust_level": item.trust_level,
            "language": source_meta.get("language", ""),
            "author": source_meta.get("author", ""),
            "date": source_meta.get("date", TODAY),
            "tags": item.tags,
            "usage_rule": item.usage_rule,
            "approved_status": "reference",
            "knowledge_tier": "reference",
            "knowledge_type": "second_layer_reference",
            "reference_date": TODAY,
            "reference_source_path": item.source_path,
            "rag_collection": COLLECTION,
        }
        body = (
            "# 核心摘要 (Summary)\n"
            f"{item.summary}\n\n"
            "# 正文内容 (Content)\n"
            f"> second_layer_reference: {item.source_path}\n\n"
            + strip_frontmatter(raw)
        )
        destination = reference_destination(item, source)
        write_markdown(destination, metadata, body)
        written.append(destination)
    return written


def extract_section(markdown: str, start_heading: str, end_heading: str) -> str:
    lines = markdown.splitlines()
    start = None
    end = None
    for index, line in enumerate(lines):
        if line.strip().lower() == start_heading.lower():
            start = index
            break
    if start is None:
        raise ValueError(f"Start heading not found: {start_heading}")
    for index in range(start + 1, len(lines)):
        if lines[index].strip().lower() == end_heading.lower():
            end = index
            break
    if end is None:
        raise ValueError(f"End heading not found: {end_heading}")
    return "\n".join(lines[start:end]).strip()


def prepare_bike_chapter() -> Path:
    source = (
        KNOWLEDGE
        / "00_inbox/processed/Triathlete Magazines Complete Triathlon Book The Training, Diet, Health, Equipment, and Safety Tips You Need to Do Your Best (Matt Fitzgerald)/ocr/Triathlete Magazines Complete Triathlon Book The Training, Diet, Health, Equipment, and Safety Tips You Need to Do Your Best (Matt Fitzgerald).md"
    )
    raw = source.read_text(encoding="utf-8", errors="ignore")
    section = extract_section(raw, "## Bike Training", "## Run Training")
    section = re.sub(r"(## Bike Training\n\n)ycling\b", r"\1Cycling", section, count=1)
    metadata = {
        "title": "Triathlete Magazine's Complete Triathlon Book - Bike Training",
        "domain": "bike",
        "source": "triathlon-knowledge/00_inbox/Triathlete Magazines Complete Triathlon Book The Training, Diet, Health, Equipment, and Safety Tips You Need to Do Your Best (Matt Fitzgerald).pdf",
        "trust_level": "B",
        "language": "en",
        "author": "Matt Fitzgerald",
        "date": TODAY,
        "tags": [
            "bike",
            "cycling",
            "bike_training",
            "cadence",
            "threshold",
            "vo2max",
            "brick",
            "periodization",
            "triathlon",
        ],
        "usage_rule": "作为骑行训练结构、骑行课类型和铁三骑跑衔接的 approved 参考；遇到功率处方时需结合 FTP、近期负荷、恢复和训练阶段。",
        "approved_status": "approved",
        "approved_date": TODAY,
        "knowledge_tier": "approved",
        "knowledge_type": "curated_bike_chapter",
        "approved_source_path": rel(source),
        "rag_collection": COLLECTION,
    }
    body = (
        "# 核心摘要 (Summary)\n"
        "这是从旧第二层 OCR 资料中抽出的骑行训练专章，补足第一层 approved 缺少 bike 独立领域资料的问题。"
        "内容覆盖骑行训练性质、踏频、技术、强度控制、恢复骑、耐力骑、长骑、阈值骑、爬坡、冲刺、brick 和骑行周期化。\n\n"
        "# 正文内容 (Content)\n"
        f"> curated_from: {rel(source)}\n\n"
        + section
    )
    destination = APPROVED / "bike" / "triathlete_magazine_complete_triathlon_book_bike_training.md"
    write_markdown(destination, metadata, body)
    return destination


def cell(ws: Any, address: str) -> Any:
    value = ws[address].value
    return "" if value is None else value


def prepare_calorie_card() -> Path:
    source = INBOX / "热量计算.xlsx"
    formula_wb = load_workbook(source, data_only=False)
    value_wb = load_workbook(source, data_only=True)
    formula_ws = formula_wb["碳水计算"]
    value_ws = value_wb["碳水计算"]
    formula_cells = [
        "D17",
        "E17",
        "F17",
        "G17",
        "H17",
        "K17",
        "L17",
        "B36",
        "M17",
        "F59",
        "N17",
        "P17",
        "Q17",
        "R17",
        "S17",
    ]
    formulas = "\n".join(f"- `{address}`: `{cell(formula_ws, address)}`" for address in formula_cells)
    example_values = "\n".join(
        [
            f"- FTP: `{cell(value_ws, 'B11')}` W",
            f"- 运动时长: `{cell(value_ws, 'B12')}` min",
            f"- 平均功率: `{cell(value_ws, 'B13')}` W",
            f"- IF: `{cell(value_ws, 'D17'):.3f}`",
            f"- 每小时机械功: `{cell(value_ws, 'G17'):.1f}` kJ/h",
            f"- 累计机械功: `{cell(value_ws, 'G17') * cell(value_ws, 'B12') / 60:.1f}` kJ",
            f"- 总能量: `{cell(value_ws, 'L17'):.1f}` kcal",
            f"- 碳水比例: `{cell(value_ws, 'N17'):.1%}`",
            f"- 碳水总量: `{cell(value_ws, 'P17'):.1f}` g",
            f"- 碳水速率: `{cell(value_ws, 'Q17'):.2f}` g/min",
            f"- 脂肪总量: `{cell(value_ws, 'R17'):.1f}` g",
            f"- 脂肪速率: `{cell(value_ws, 'S17'):.2f}` g/min",
        ]
    )
    metadata = {
        "title": "骑行热量与碳水消耗计算模板",
        "domain": "bike",
        "source": rel(source),
        "trust_level": "B",
        "language": "zh",
        "author": "",
        "date": TODAY,
        "tags": [
            "bike",
            "cycling_power",
            "FTP",
            "IF",
            "carbohydrate",
            "energy_expenditure",
            "nutrition_planning",
        ],
        "usage_rule": "用于根据 FTP、时长和平均功率估算骑行能量消耗、碳水消耗和脂肪消耗；只适合 IF<=1 且 IF>0.55 的相对稳态骑行，不可替代实验室代谢测试。",
        "approved_status": "approved",
        "approved_date": TODAY,
        "knowledge_tier": "approved",
        "knowledge_type": "calculator_template",
        "approved_source_path": rel(source),
        "rag_collection": COLLECTION,
    }
    body = f"""# 核心摘要 (Summary)
这是一张用于后续课表和补给安排的骑行能量计算卡。它把 Excel 里的关键输入、计算链和适用边界转成 RAG 可检索文本：输入 FTP、运动时长、平均功率和总效率，输出机械功、总能量、碳水比例、碳水克数、碳水每分钟速率、脂肪克数和脂肪每分钟速率。

# 正文内容 (Content)
## 推荐入库形态

这份资料不适合只作为普通长文 reference。它应该作为 `calculator_template` 放在 `01_approved/bike`，因为后续问答器或课表生成器可以先检索到这张卡，再调用真正的计算函数复现公式。

## 输入字段

- FTP: 单位 W，对应 Excel `B11`。
- 运动时长: 单位分钟，对应 Excel `B12`。
- 平均功率: 单位 W，对应 Excel `B13`。
- 总效率: 默认 `{cell(value_ws, 'E17')}`，对应 Excel `E17`。

## 计算链

{formulas}

## 当前表格示例

{example_values}

## 使用边界

- 适用于相对稳态的骑行训练、长骑、功率车训练和大铁骑行补给估算。
- 表格注释指出：该模型只适合 IF 小于等于 1 且大于 0.55 的情况。
- QR 和底物比例是估算值，会受饮食、训练状态、持续时间、环境和个体差异影响。
- 回答补给问题时，不能只凭这张表直接开完整补给处方，还要追问体重、出汗率、胃肠耐受、训练环境和既往补给实践。
"""
    destination = APPROVED / "bike" / "cycling_energy_carbohydrate_calculator.md"
    write_markdown(destination, metadata, body)
    return destination


def prepare_periodization_card() -> Path:
    source = INBOX / "周期训练计划模板.xlsx"
    wb = load_workbook(source, data_only=False)
    ws = wb["Sheet1"]
    metadata = {
        "title": "骑行功率周期训练计划模板",
        "domain": "bike",
        "source": rel(source),
        "trust_level": "B",
        "language": "zh",
        "author": "",
        "date": TODAY,
        "tags": [
            "bike",
            "cycling_power",
            "periodization",
            "macrocycle",
            "mesocycle",
            "microcycle",
            "volume",
            "intensity",
            "workload",
            "lesson_planning",
        ],
        "usage_rule": "用于把骑行功率训练拆成年度、月周期、周周期、训练阶段、容量、强度和负荷变化；生成课表前必须结合目标赛事日期、当前 FTP、CTL/ATL/TSB、可训练时间和恢复情况。",
        "approved_status": "approved",
        "approved_date": TODAY,
        "knowledge_tier": "approved",
        "knowledge_type": "periodization_template",
        "approved_source_path": rel(source),
        "rag_collection": COLLECTION,
    }
    phase_rows = "\n".join(
        [
            "- `Macrocycle`: 年度或大周期层级。",
            "- `Mesocycle`: 月周期或阶段周期层级。",
            "- `Microcycle`: 周周期层级。",
            "- `Hypertrophy / Strength / Str-Power / Power / Peaking`: 力量和功率发展阶段。",
            "- `Volume 1-10`: 训练容量等级。",
            "- `Intensity 1-10`: 训练强度等级。",
            "- `WorkLoad`: `Load+`、`Load`、`Base`、`Deload` 等负荷状态。",
        ]
    )
    body = f"""# 核心摘要 (Summary)
这是一张用于后续安排骑行功率训练课表的周期化模板卡。它不是普通参考书，而是 `periodization_template`：RAG 负责检索模板结构，后续应用层负责把目标赛事、FTP、可训练时间和疲劳状态填进去。

# 正文内容 (Content)
## 推荐入库形态

这份 Excel 更适合转成 `01_approved/bike` 下的模板说明卡，而不是整表逐格切 chunk。原因是它的价值在结构：年度起始日期、周序列、宏周期、月周期、周周期、阶段、容量、强度和负荷，而不是每个空白单元格。

## 表格结构

- 工作表: `{ws.title}`。
- 队伍名称字段: `I4` / `L4`，当前示例值为 `{cell(ws, 'L4')}`。
- 年度开始日期字段: `Q4` / `V4`，当前示例值为 `{cell(ws, 'V4')}`。
- 日期列从 `F` 到 `BE` 展开，要求年度开始日期是周一，方便按周排课。

## 周期化字段

{phase_rows}

## RAG 使用方式

- 当用户问“怎么安排骑行功率周期训练”“怎么排接下来 8-12 周骑行课表”“功率训练怎么做大周期/小周期”时，应优先召回这张模板卡。
- RAG 只回答模板结构和排课依据；真正排课前必须追问或读取目标赛事日期、当前 FTP、最近四到六周训练量、关键约束、每周可训练天数、伤病和疲劳状态。
- 这张模板适合和 `The Triathlete's Training Bible` 的周期化原则、`Triathlon Science` 的训练负荷概念、以及骑行热量计算模板一起使用。

## 应用层建议

后续应用最好把这张模板拆成结构化对象：

```json
{{
  "macrocycle": "year_or_season",
  "mesocycle": "block",
  "microcycle": "week",
  "week_fields": ["date", "competition", "primary_goal", "secondary_goal"],
  "load_fields": ["volume_1_10", "intensity_1_10", "workload_state"],
  "phase_fields": ["hypertrophy", "strength", "strength_power", "power", "peaking"]
}}
```
"""
    destination = APPROVED / "bike" / "cycling_power_periodization_template.md"
    write_markdown(destination, metadata, body)
    return destination


def main() -> int:
    written = []
    written.extend(prepare_references())
    written.append(prepare_bike_chapter())
    written.append(prepare_calorie_card())
    written.append(prepare_periodization_card())
    print(f"written={len(written)}")
    for path in written:
        print(rel(path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
